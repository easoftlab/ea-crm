#!/usr/bin/env python3
"""
Report Export System
Provides PDF, Excel, and CSV export functionality for all reports
"""

import os
import sqlite3
import json
from datetime import datetime, timedelta
from io import BytesIO
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
import csv

class ReportExporter:
    """Handles export of reports in various formats."""
    
    def __init__(self, db_path='instance/leads.db'):
        self.db_path = db_path
        self.styles = getSampleStyleSheet()
        self.setup_custom_styles()
    
    def setup_custom_styles(self):
        """Setup custom paragraph styles for PDF reports."""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=12,
            spaceAfter=20
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomBody',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=12
        ))
    
    def export_daily_report_pdf(self, user_id, date=None):
        """Export daily report as PDF."""
        if date is None:
            date = datetime.now().date()
        
        # Get report data
        report_data = self.get_daily_report_data(user_id, date)
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Title
        story.append(Paragraph(f"Daily Report - {date.strftime('%B %d, %Y')}", self.styles['CustomTitle']))
        story.append(Spacer(1, 20))
        
        # User info
        story.append(Paragraph(f"Team Member: {report_data['username']}", self.styles['CustomSubtitle']))
        story.append(Paragraph(f"Team: {report_data['team_name']}", self.styles['CustomBody']))
        story.append(Spacer(1, 20))
        
        # Key metrics table
        metrics_data = [
            ['Metric', 'Value', 'Goal', 'Achievement %'],
            ['Leads Created', str(report_data['leads_created']), '10', f"{report_data['leads_created']/10*100:.1f}%"],
            ['Tasks Completed', str(report_data['tasks_completed']), '5', f"{report_data['tasks_completed']/5*100:.1f}%"],
            ['Messages Sent', str(report_data['messages_sent']), '20', f"{report_data['messages_sent']/20*100:.1f}%"],
            ['Productivity Score', f"{report_data['productivity_score']:.1f}", '100', f"{report_data['productivity_score']:.1f}%"]
        ]
        
        metrics_table = Table(metrics_data)
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(metrics_table)
        story.append(Spacer(1, 20))
        
        # Notes section
        if report_data.get('notes'):
            story.append(Paragraph("Notes:", self.styles['CustomSubtitle']))
            story.append(Paragraph(report_data['notes'], self.styles['CustomBody']))
            story.append(Spacer(1, 20))
        
        # Manager notes section
        if report_data.get('manager_notes'):
            story.append(Paragraph("Manager Notes:", self.styles['CustomSubtitle']))
            story.append(Paragraph(report_data['manager_notes'], self.styles['CustomBody']))
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def export_weekly_report_pdf(self, user_id, week_start=None):
        """Export weekly report as PDF."""
        if week_start is None:
            week_start = datetime.now().date() - timedelta(days=datetime.now().weekday())
        
        # Get report data
        report_data = self.get_weekly_report_data(user_id, week_start)
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Title
        week_end = week_start + timedelta(days=6)
        story.append(Paragraph(f"Weekly Report - {week_start.strftime('%B %d')} to {week_end.strftime('%B %d, %Y')}", self.styles['CustomTitle']))
        story.append(Spacer(1, 20))
        
        # User info
        story.append(Paragraph(f"Team Member: {report_data['username']}", self.styles['CustomSubtitle']))
        story.append(Paragraph(f"Team: {report_data['team_name']}", self.styles['CustomBody']))
        story.append(Spacer(1, 20))
        
        # Weekly metrics table
        metrics_data = [
            ['Metric', 'This Week', 'Last Week', 'Change %'],
            ['Total Leads', str(report_data['total_leads_created']), str(report_data['last_week_leads']), f"{report_data['leads_change']:.1f}%"],
            ['Tasks Completed', str(report_data['total_tasks_completed']), str(report_data['last_week_tasks']), f"{report_data['tasks_change']:.1f}%"],
            ['Messages Sent', str(report_data['total_messages_sent']), str(report_data['last_week_messages']), f"{report_data['messages_change']:.1f}%"],
            ['Productivity Score', f"{report_data['productivity_score']:.1f}", f"{report_data['last_week_productivity']:.1f}", f"{report_data['productivity_change']:.1f}%"]
        ]
        
        metrics_table = Table(metrics_data)
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(metrics_table)
        story.append(Spacer(1, 20))
        
        # Performance trends
        story.append(Paragraph("Performance Trends:", self.styles['CustomSubtitle']))
        story.append(Paragraph(f"Trend: {report_data['productivity_trend']}", self.styles['CustomBody']))
        story.append(Paragraph(f"Performance Rating: {report_data['performance_rating']}", self.styles['CustomBody']))
        story.append(Paragraph(f"Workload Distribution: {report_data['workload_distribution']}", self.styles['CustomBody']))
        story.append(Spacer(1, 20))
        
        # Notes
        if report_data.get('notes'):
            story.append(Paragraph("Notes:", self.styles['CustomSubtitle']))
            story.append(Paragraph(report_data['notes'], self.styles['CustomBody']))
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def export_monthly_report_pdf(self, user_id, month_year=None):
        """Export monthly report as PDF."""
        if month_year is None:
            month_year = datetime.now().strftime('%Y-%m')
        
        # Get report data
        report_data = self.get_monthly_report_data(user_id, month_year)
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Title
        month_name = datetime.strptime(month_year, '%Y-%m').strftime('%B %Y')
        story.append(Paragraph(f"Monthly Report - {month_name}", self.styles['CustomTitle']))
        story.append(Spacer(1, 20))
        
        # User info
        story.append(Paragraph(f"Team Member: {report_data['username']}", self.styles['CustomSubtitle']))
        story.append(Paragraph(f"Team: {report_data['team_name']}", self.styles['CustomBody']))
        story.append(Spacer(1, 20))
        
        # Monthly metrics table
        metrics_data = [
            ['Metric', 'This Month', 'Last Month', 'Change %'],
            ['Total Leads', str(report_data['total_leads_created']), str(report_data['last_month_leads']), f"{report_data['leads_change']:.1f}%"],
            ['Tasks Completed', str(report_data['total_tasks_completed']), str(report_data['last_month_tasks']), f"{report_data['tasks_change']:.1f}%"],
            ['Messages Sent', str(report_data['total_messages_sent']), str(report_data['last_month_messages']), f"{report_data['messages_change']:.1f}%"],
            ['Productivity Score', f"{report_data['productivity_score']:.1f}", f"{report_data['last_month_productivity']:.1f}", f"{report_data['productivity_change']:.1f}%"]
        ]
        
        metrics_table = Table(metrics_data)
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(metrics_table)
        story.append(Spacer(1, 20))
        
        # Team ranking
        story.append(Paragraph("Team Performance:", self.styles['CustomSubtitle']))
        story.append(Paragraph(f"Team Rank: {report_data['team_rank']}", self.styles['CustomBody']))
        story.append(Paragraph(f"Performance Percentile: {report_data['team_performance_percentile']:.1f}%", self.styles['CustomBody']))
        story.append(Paragraph(f"Improvement from Last Month: {report_data['improvement_from_last_month']:.1f}%", self.styles['CustomBody']))
        story.append(Spacer(1, 20))
        
        # Key achievements
        if report_data.get('key_achievements'):
            story.append(Paragraph("Key Achievements:", self.styles['CustomSubtitle']))
            story.append(Paragraph(report_data['key_achievements'], self.styles['CustomBody']))
            story.append(Spacer(1, 20))
        
        # Areas for improvement
        if report_data.get('areas_for_improvement'):
            story.append(Paragraph("Areas for Improvement:", self.styles['CustomSubtitle']))
            story.append(Paragraph(report_data['areas_for_improvement'], self.styles['CustomBody']))
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def export_team_report_excel(self, team_id, report_type='daily', date=None):
        """Export team report as Excel file."""
        if date is None:
            date = datetime.now().date()
        
        # Get team report data
        team_data = self.get_team_report_data(team_id, report_type, date)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_type.title()} Report"
        
        # Header
        ws['A1'] = f"{report_type.title()} Team Report - {date.strftime('%B %d, %Y')}"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')
        
        # Team info
        ws['A3'] = f"Team: {team_data['team_name']}"
        ws['A4'] = f"Total Members: {len(team_data['members'])}"
        ws['A5'] = f"Active Members: {team_data['active_members']}"
        
        # Headers
        headers = ['Member', 'Leads Created', 'Tasks Completed', 'Messages Sent', 'Productivity Score']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF")
        
        # Data rows
        for row, member in enumerate(team_data['members'], 8):
            ws.cell(row=row, column=1, value=member['username'])
            ws.cell(row=row, column=2, value=member['leads_created'])
            ws.cell(row=row, column=3, value=member['tasks_completed'])
            ws.cell(row=row, column=4, value=member['messages_sent'])
            ws.cell(row=row, column=5, value=member['productivity_score'])
        
        # Summary row
        summary_row = len(team_data['members']) + 8
        ws.cell(row=summary_row, column=1, value="TOTAL")
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        
        for col in range(2, 6):
            cell = ws.cell(row=summary_row, column=col)
            cell.value = f"=SUM({chr(64+col)}8:{chr(64+col)}{summary_row-1})"
            cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def export_report_csv(self, report_type, filters=None):
        """Export report data as CSV."""
        if filters is None:
            filters = {}
        
        # Get report data based on type
        if report_type == 'daily':
            data = self.get_all_daily_reports(filters)
        elif report_type == 'weekly':
            data = self.get_all_weekly_reports(filters)
        elif report_type == 'monthly':
            data = self.get_all_monthly_reports(filters)
        else:
            raise ValueError(f"Unknown report type: {report_type}")
        
        # Create CSV buffer
        buffer = BytesIO()
        writer = csv.writer(buffer)
        
        # Write headers
        if data:
            writer.writerow(data[0].keys())
            writer.writerows([row.values() for row in data])
        
        buffer.seek(0)
        return buffer
    
    def get_daily_report_data(self, user_id, date):
        """Get daily report data for a user."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT dr.*, u.username, t.name as team_name
            FROM team_member_daily_reports dr
            JOIN users u ON dr.user_id = u.id
            JOIN teams t ON dr.team_id = t.id
            WHERE dr.user_id = ? AND dr.report_date = ?
        """, (user_id, date.isoformat()))
        
        row = cursor.fetchone()
        conn.close()
        
        if row:
            return {
                'username': row[1],
                'team_name': row[2],
                'leads_created': row[3],
                'tasks_completed': row[4],
                'messages_sent': row[5],
                'productivity_score': row[6],
                'notes': row[7],
                'manager_notes': row[8]
            }
        return {}
    
    def get_weekly_report_data(self, user_id, week_start):
        """Get weekly report data for a user."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT wr.*, u.username, t.name as team_name
            FROM team_member_weekly_reports wr
            JOIN users u ON wr.user_id = u.id
            JOIN teams t ON wr.team_id = t.id
            WHERE wr.user_id = ? AND wr.week_start = ?
        """, (user_id, week_start.isoformat()))
        
        row = cursor.fetchone()
        conn.close()
        
        if row:
            return {
                'username': row[1],
                'team_name': row[2],
                'total_leads_created': row[3],
                'total_tasks_completed': row[4],
                'total_messages_sent': row[5],
                'productivity_score': row[6],
                'productivity_trend': row[7],
                'performance_rating': row[8],
                'workload_distribution': row[9],
                'notes': row[10],
                'last_week_leads': 0,  # Would need to calculate
                'last_week_tasks': 0,
                'last_week_messages': 0,
                'last_week_productivity': 0,
                'leads_change': 0,
                'tasks_change': 0,
                'messages_change': 0,
                'productivity_change': 0
            }
        return {}
    
    def get_monthly_report_data(self, user_id, month_year):
        """Get monthly report data for a user."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT mr.*, u.username, t.name as team_name
            FROM team_member_monthly_reports mr
            JOIN users u ON mr.user_id = u.id
            JOIN teams t ON mr.team_id = t.id
            WHERE mr.user_id = ? AND mr.month_year = ?
        """, (user_id, month_year))
        
        row = cursor.fetchone()
        conn.close()
        
        if row:
            return {
                'username': row[1],
                'team_name': row[2],
                'total_leads_created': row[3],
                'total_tasks_completed': row[4],
                'total_messages_sent': row[5],
                'productivity_score': row[6],
                'team_rank': row[7],
                'team_performance_percentile': row[8],
                'improvement_from_last_month': row[9],
                'key_achievements': row[10],
                'areas_for_improvement': row[11],
                'last_month_leads': 0,  # Would need to calculate
                'last_month_tasks': 0,
                'last_month_messages': 0,
                'last_month_productivity': 0,
                'leads_change': 0,
                'tasks_change': 0,
                'messages_change': 0,
                'productivity_change': 0
            }
        return {}
    
    def get_team_report_data(self, team_id, report_type, date):
        """Get team report data."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Get team info
        cursor.execute("SELECT name FROM teams WHERE id = ?", (team_id,))
        team_name = cursor.fetchone()[0]
        
        # Get team members with their report data
        if report_type == 'daily':
            cursor.execute("""
                SELECT u.username, dr.leads_created, dr.tasks_completed, dr.messages_sent, dr.productivity_score
                FROM users u
                JOIN user_teams ut ON u.id = ut.user_id
                LEFT JOIN team_member_daily_reports dr ON u.id = dr.user_id AND dr.report_date = ?
                WHERE ut.team_id = ?
            """, (date.isoformat(), team_id))
        elif report_type == 'weekly':
            week_start = date - timedelta(days=date.weekday())
            cursor.execute("""
                SELECT u.username, wr.total_leads_created, wr.total_tasks_completed, wr.total_messages_sent, wr.productivity_score
                FROM users u
                JOIN user_teams ut ON u.id = ut.user_id
                LEFT JOIN team_member_weekly_reports wr ON u.id = wr.user_id AND wr.week_start = ?
                WHERE ut.team_id = ?
            """, (week_start.isoformat(), team_id))
        else:
            cursor.execute("""
                SELECT u.username, 0, 0, 0, 0
                FROM users u
                JOIN user_teams ut ON u.id = ut.user_id
                WHERE ut.team_id = ?
            """, (team_id,))
        
        members = []
        active_members = 0
        
        for row in cursor.fetchall():
            member = {
                'username': row[0],
                'leads_created': row[1] or 0,
                'tasks_completed': row[2] or 0,
                'messages_sent': row[3] or 0,
                'productivity_score': row[4] or 0
            }
            members.append(member)
            
            if member['productivity_score'] > 0:
                active_members += 1
        
        conn.close()
        
        return {
            'team_name': team_name,
            'members': members,
            'active_members': active_members
        }
    
    def get_all_daily_reports(self, filters):
        """Get all daily reports with filters."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = """
            SELECT dr.*, u.username, t.name as team_name
            FROM team_member_daily_reports dr
            JOIN users u ON dr.user_id = u.id
            JOIN teams t ON dr.team_id = t.id
        """
        
        params = []
        conditions = []
        
        if 'team_id' in filters:
            conditions.append("dr.team_id = ?")
            params.append(filters['team_id'])
        
        if 'date_from' in filters:
            conditions.append("dr.report_date >= ?")
            params.append(filters['date_from'])
        
        if 'date_to' in filters:
            conditions.append("dr.report_date <= ?")
            params.append(filters['date_to'])
        
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        
        # Convert to list of dictionaries
        columns = [description[0] for description in cursor.description]
        return [dict(zip(columns, row)) for row in rows]
    
    def get_all_weekly_reports(self, filters):
        """Get all weekly reports with filters."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = """
            SELECT wr.*, u.username, t.name as team_name
            FROM team_member_weekly_reports wr
            JOIN users u ON wr.user_id = u.id
            JOIN teams t ON wr.team_id = t.id
        """
        
        params = []
        conditions = []
        
        if 'team_id' in filters:
            conditions.append("wr.team_id = ?")
            params.append(filters['team_id'])
        
        if 'week_from' in filters:
            conditions.append("wr.week_start >= ?")
            params.append(filters['week_from'])
        
        if 'week_to' in filters:
            conditions.append("wr.week_start <= ?")
            params.append(filters['week_to'])
        
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        
        # Convert to list of dictionaries
        columns = [description[0] for description in cursor.description]
        return [dict(zip(columns, row)) for row in rows]
    
    def get_all_monthly_reports(self, filters):
        """Get all monthly reports with filters."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = """
            SELECT mr.*, u.username, t.name as team_name
            FROM team_member_monthly_reports mr
            JOIN users u ON mr.user_id = u.id
            JOIN teams t ON mr.team_id = t.id
        """
        
        params = []
        conditions = []
        
        if 'team_id' in filters:
            conditions.append("mr.team_id = ?")
            params.append(filters['team_id'])
        
        if 'month_from' in filters:
            conditions.append("mr.month_year >= ?")
            params.append(filters['month_from'])
        
        if 'month_to' in filters:
            conditions.append("mr.month_year <= ?")
            params.append(filters['month_to'])
        
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        
        # Convert to list of dictionaries
        columns = [description[0] for description in cursor.description]
        return [dict(zip(columns, row)) for row in rows]

# Example usage
if __name__ == "__main__":
    exporter = ReportExporter()
    
    # Export daily report as PDF
    pdf_buffer = exporter.export_daily_report_pdf(1, datetime.now().date())
    
    # Export team report as Excel
    excel_buffer = exporter.export_team_report_excel(1, 'daily', datetime.now().date())
    
    # Export all daily reports as CSV
    csv_buffer = exporter.export_report_csv('daily', {'team_id': 1})
    
    print("Export completed successfully!") 